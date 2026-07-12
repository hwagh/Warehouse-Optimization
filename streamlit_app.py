"""
Warehouse Capacity Planner — Streamlit Web App
Run:  streamlit run streamlit_app.py
"""

import streamlit as st
import plotly.graph_objects as go
import plotly.express as px
import pandas as pd
import copy, sys, os, json

sys.path.insert(0, os.path.dirname(__file__))

from config import (
    StorageArea, OrderType,
    StorageSplit, CustomerSplit, KittingSplit, KitStorageSplit, KitSourceSplit,
    ZONE_NAMES, ZONE_FLOW_ORDER,
    DEFAULT_AREAS, DEFAULT_ORDER_TYPES,
)
from engine import WarehouseEngine
import database as db

st.set_page_config(
    page_title="Warehouse Capacity Planner",
    page_icon="⬡", layout="wide",
    initial_sidebar_state="expanded",
)

# ── Global styling — spacing, sizing, and visual polish applied to every page ──
# Applied once to every page.
st.markdown("""
<style>
/* hide only the GitHub source-code icon in the hosted-app toolbar
   (leaves Share, star, edit, and the menu visible) */
[data-testid="stToolbarActions"] a[href*="github.com"] { display: none !important; }
/* reclaim the big empty margin at the top + sides of the page */
.block-container, [data-testid="stMainBlockContainer"] {
    padding-top: 3.2rem !important;
    padding-bottom: 1rem !important;
    padding-left: 1.6rem !important;
    padding-right: 1.6rem !important;
    max-width: 100% !important;
}
/* tighten the vertical gap between stacked elements */
[data-testid="stVerticalBlock"] { gap: 0.6rem !important; }
[data-testid="stHorizontalBlock"] { gap: 0.65rem !important; }
/* headings — larger, clearer hierarchy */
h1, [data-testid="stHeading"] h1 { font-size: 2.35rem !important; margin: .1rem 0 .55rem 0 !important; }
h2 { font-size: 1.7rem  !important; margin: .4rem 0 !important; padding: 0 !important; }
h3 { font-size: 1.35rem !important; margin: .32rem 0 !important; padding: 0 !important; }
/* base body text */
[data-testid="stMarkdownContainer"], [data-testid="stMarkdownContainer"] p,
.stMarkdown, .stMarkdown li { font-size: 1.12rem !important; }
/* captions + small text */
[data-testid="stCaptionContainer"], .stCaption { font-size: 1.0rem !important; margin: 0 !important; }
/* markdown paragraph spacing */
[data-testid="stMarkdownContainer"] p { margin-bottom: .35rem !important; line-height: 1.55 !important; }
/* thin dividers */
hr { margin: .55rem 0 !important; }
/* metrics — bigger value, readable label */
[data-testid="stMetric"] { padding: .4rem .6rem !important; }
[data-testid="stMetricValue"] { font-size: 1.9rem !important; }
[data-testid="stMetricLabel"] { font-size: 1.02rem !important; }
[data-testid="stMetricLabel"] p { font-size: 1.02rem !important; }
/* dataframes / data_editor: readable font + rows */
[data-testid="stDataFrame"], [data-testid="stDataEditor"] { font-size: 1.05rem !important; }
[data-testid="stDataFrame"] div, [data-testid="stDataEditor"] div { line-height: 1.4 !important; }
/* expanders: readable header + body */
[data-testid="stExpander"] summary { padding: .5rem .75rem !important; font-size: 1.15rem !important; }
[data-testid="stExpander"] [data-testid="stExpanderDetails"] { padding: .5rem .8rem !important; }
/* buttons: comfortable padding + text */
.stButton button, [data-testid="stDownloadButton"] button, [data-testid="baseButton-primary"] {
    padding: .5rem 1rem !important; min-height: 0 !important; font-size: 1.08rem !important;
}
/* alert / info / success boxes */
[data-testid="stAlert"], [data-testid="stAlertContainer"] { padding: .65rem .85rem !important; }
[data-testid="stAlert"] p { margin: 0 !important; font-size: 1.08rem !important; }
/* number inputs a touch shorter */
[data-testid="stNumberInput"] input { padding-top: .35rem !important; padding-bottom: .35rem !important; font-size: 1.08rem !important; }
/* tabs: readable tab bar */
[data-testid="stTabs"] [data-baseweb="tab"] { padding: .55rem 1rem !important; }
[data-testid="stTabs"] [data-baseweb="tab"] p { font-size: 1.15rem !important; }
[data-testid="stTabs"] [data-baseweb="tab-list"] { gap: .35rem !important; }
/* sidebar padding */
[data-testid="stSidebar"] [data-testid="stVerticalBlock"] { gap: .45rem !important; }
[data-testid="stSidebar"] .block-container { padding-top: 1rem !important; }
/* sidebar nav radio — larger, easier to click */
[data-testid="stSidebar"] [data-testid="stRadio"] label p { font-size: 1.08rem !important; }

/* ── visual polish: consistent grouping, hierarchy, accents ── */
/* KPI metrics read as a row of aligned cards */
[data-testid="stMetric"] {
    background: rgba(255,255,255,.03);
    border: 1px solid rgba(255,255,255,.08);
    border-radius: 8px;
}
[data-testid="stMetricValue"] { font-weight: 700 !important; }
[data-testid="stMetricLabel"] { opacity: .72; }
/* bordered containers (getting-started, groups) softer + consistent */
[data-testid="stVerticalBlockBorderWrapper"] {
    border-radius: 10px !important;
    border-color: rgba(255,255,255,.08) !important;
}
/* active tab gets the brand accent */
[data-testid="stTabs"] [data-baseweb="tab"][aria-selected="true"] { color: #8ea2ff !important; }
[data-testid="stTabs"] [data-baseweb="tab-highlight"] { background-color: #4f6ef7 !important; }
/* subheaders: subtle accent underline for clearer hierarchy */
h3 { border-bottom: 1px solid rgba(255,255,255,.06); padding-bottom: .12rem !important; }
/* expander header hover affordance */
[data-testid="stExpander"] summary:hover { color: #8ea2ff !important; }
</style>
""", unsafe_allow_html=True)

ZONE_COLORS = {
    "600": "#4f6ef7",
    "400": "#06b6d4", "300": "#10b981",
    "200": "#f59e0b", "100": "#ef4444",
}

def status_color(pct):
    if pct >= 100: return "#ef4444"
    if pct >= 85:  return "#f59e0b"
    if pct >= 70:  return "#f97316"
    return "#22c55e"

def status_label(pct):
    if pct >= 100: return "🔴 OVER CAPACITY"
    if pct >= 85:  return "🟡 CRITICAL"
    if pct >= 70:  return "🟠 WARNING"
    return "🟢 OK"


def status_label_for(area_snapshot):
    """Overflow-aware status label using the area's own cascade logic."""
    s = area_snapshot.status  # already computed with overflow in the engine
    return {
        "OVER CAPACITY": "🔴 OVER CAPACITY",
        "CRITICAL":      "🟡 CRITICAL",
        "WARNING":       "🟠 WARNING",
        "OK":            "🟢 OK",
    }.get(s, "🟢 OK")


def render_last_updated():
    """Show when the shared setup was last saved. Same value for every viewer,
    since it's read from stored data (Supabase or the local JSON), not the
    current session clock."""
    from datetime import datetime as _dt
    ts = None
    try:
        ts = db.last_updated()
    except Exception:
        ts = None
    if not ts:
        st.caption("🕒 Last updated: not yet saved")
        return
    try:
        dt = _dt.fromisoformat(ts)
        pretty = dt.strftime("%b %d, %Y at %I:%M %p") + " UTC"
    except Exception:
        pretty = str(ts)
    st.caption("🕒 Last updated: " + pretty)


# ── unit system ───────────────────────────────────────────────────────────────
# Internal storage is always cubic feet.
# All inputs/outputs convert to/from the selected display unit.

UNITS = {
    "cu ft":  {"label": "cu ft",  "symbol": "ft³",  "to_cuft": 1.0,          "from_cuft": 1.0},
    "cu in":  {"label": "cu in",  "symbol": "in³",  "to_cuft": 1/1728,       "from_cuft": 1728.0},
    "cu cm":  {"label": "cu cm",  "symbol": "cm³",  "to_cuft": 1/28316.8,    "from_cuft": 28316.8},
    "cu m":   {"label": "cu m",   "symbol": "m³",   "to_cuft": 35.3147,      "from_cuft": 1/35.3147},
}

def to_display(cuft_value: float) -> float:
    """Convert internal cu ft value to display unit."""
    unit = st.session_state.get("display_unit", "cu ft")
    return cuft_value * UNITS[unit]["from_cuft"]

def to_cuft(display_value: float) -> float:
    """Convert display unit value to internal cu ft."""
    unit = st.session_state.get("display_unit", "cu ft")
    return display_value * UNITS[unit]["to_cuft"]

def unit_label() -> str:
    unit = st.session_state.get("display_unit", "cu ft")
    return UNITS[unit]["symbol"]

def fmt_vol(cuft_value: float, decimals: int = 1) -> str:
    """Format a volume in the current display unit."""
    return f"{to_display(cuft_value):,.{decimals}f} {unit_label()}"


# ── linear dimension helpers ─────────────────────────────────────────────────
# Rack/box L·D·H are stored internally in FEET (volume = L×D×H gives cubic feet).
# The reference data was in centimetres and converted once to feet (feet =
# cm / 30.48); everything on screen is shown in feet.
def fmt_dims_ft(l_ft: float, d_ft: float, h_ft: float, decimals: int = 2) -> str:
    """Format an L×D×H triple (already stored in feet) as feet."""
    return "×".join(f"{v:.{decimals}f}" for v in (l_ft, d_ft, h_ft))


def _config_signature() -> str:
    """Fingerprint of the current config at the *editable* precision.

    Dimensions are compared as feet rounded to 2 dp (what the grid shows), so
    float noise never counts as a real edit — only a value a user actually
    changed does. Used to drive auto-save.
    """
    def area_key(a):
        g = lambda k: getattr(a, k, 0) or 0
        return [
            a.id, a.name, a.zone,
            round(a.rack_length_cuft, 2), round(a.rack_depth_cuft, 2),
            round(a.rack_height_cuft, 2), int(a.num_racks),
            round(a.box_length_cuft, 2), round(a.box_depth_cuft, 2),
            round(a.box_height_cuft, 2), round(a.efficiency, 4),
            round(a.units_per_box, 3),
            round(g("of1_rack_length_cuft"), 2), round(g("of1_rack_depth_cuft"), 2),
            round(g("of1_rack_height_cuft"), 2), int(g("of1_num_racks")),
            round(g("of2_rack_length_cuft"), 2), round(g("of2_rack_depth_cuft"), 2),
            round(g("of2_rack_height_cuft"), 2), int(g("of2_num_racks")),
        ]

    def ot_key(o):
        # Use getattr with defaults so order types loaded before the kit-split
        # fields existed (stale session state) never crash the signature.
        ks = getattr(o, "kit_storage_split", None)
        ksrc = getattr(o, "kit_source_split", None)
        kit300 = round(ks.kit300_pct, 3) if ks else 60.0
        kit200 = round(ks.kit200_pct, 3) if ks else 40.0
        kit600 = round(ksrc.kit600_pct, 3) if ksrc else 40.0
        kit400 = round(ksrc.kit400_pct, 3) if ksrc else 60.0
        return [
            o.id, o.name, int(o.daily_volume), int(o.avg_units_per_order),
            round(o.storage_split.paper_pct, 3), round(o.storage_split.consumable_pct, 3),
            round(o.customer_split.cust1_pct, 3), round(o.customer_split.cust2_pct, 3),
            round(o.kitting_split.packout_pct, 3), round(o.kitting_split.kitting_pct, 3),
            kit300, kit200, kit600, kit400,
        ]

    return json.dumps({
        "a": [area_key(a) for a in st.session_state.areas],
        "o": [ot_key(o) for o in st.session_state.order_types],
    }, sort_keys=True, default=str)


# ── Single-file CSV template ─────────────────────────────────────────────────
# One CSV with a SECTION column keeps everything together and stays
# human-readable — open in Excel, fill in, upload back.

import io
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

AREA_COLS = [
    ("area_id",              "Area ID",             "Unique key - do not change"),
    ("area_name",            "Area Name",           "Display name - safe to edit"),
    ("zone",                 "Zone",                "Zone code - do not change"),
    ("rack_length_cuft",     "Rack Length (ft)",    "Length of one rack in feet"),
    ("rack_depth_cuft",      "Rack Depth (ft)",     "Depth of one rack in feet"),
    ("rack_height_cuft",     "Rack Height (ft)",    "Height of one rack in feet"),
    ("num_racks",            "Number of Racks",     "How many racks in this area"),
    ("box_length_cuft",      "Box Length (ft)",     "Length of average box in feet"),
    ("box_depth_cuft",       "Box Width (ft)",      "Width of average box in feet"),
    ("box_height_cuft",      "Box Height (ft)",     "Height of average box in feet"),
    ("efficiency",           "Efficiency (0-1)",    "Usable fraction, e.g. 0.75 = 75% after aisles"),
    ("units_per_box",        "Units per Box",       "Average units that fit in one box in this area"),
    ("of1_rack_length_cuft",  "OF1 Rack Length (ft)",  "Overflow zone rack dimension (0 = none)"),
    ("of1_rack_depth_cuft",   "OF1 Rack Depth (ft)",   "Overflow zone rack dimension (0 = none)"),
    ("of1_rack_height_cuft",  "OF1 Rack Height (ft)",  "Overflow zone rack dimension (0 = none)"),
    ("of1_num_racks",         "OF1 Number of Racks",   "Overflow zone rack dimension (0 = none)"),
    ("of2_rack_length_cuft",  "OF2 Rack Length (ft)",  "Overflow zone rack dimension (0 = none)"),
    ("of2_rack_depth_cuft",   "OF2 Rack Depth (ft)",   "Overflow zone rack dimension (0 = none)"),
    ("of2_rack_height_cuft",  "OF2 Rack Height (ft)",  "Overflow zone rack dimension (0 = none)"),
    ("of2_num_racks",         "OF2 Number of Racks",   "Overflow zone rack dimension (0 = none)"),
]

ORDER_COLS = [
    ("order_id",            "Order ID",            "Unique key - do not change"),
    ("order_name",          "Order Name",          "Display name - safe to edit"),
    ("daily_volume",        "Daily Volume",        "Number of orders per day"),
    ("avg_units_per_order", "Avg Units / Order",   "Average units per order"),
    ("paper_pct",           "Paper % (600)",       "Must total 100 with Consumable %"),
    ("consumable_pct",      "Consumable % (400)",  "Must total 100 with Paper %"),
    ("cust1_pct",           "Cust 1 % (Zone 300)", "Must total 100 with Cust 2 %"),
    ("cust2_pct",           "Cust 2 % (Zone 200)", "Must total 100 with Cust 1 %"),
    ("packout_pct",         "Direct to Packout %", "Must total 100 with Kitting %"),
    ("kitting_pct",         "Kitting %",           "Must total 100 with Direct to Packout %"),
    ("kit300_pct",          "Kit 300 %",           "Custom-kit storage: must total 100 with Kit 200 %"),
    ("kit200_pct",          "Kit 200 %",           "Custom-kit storage: must total 100 with Kit 300 %"),
    ("kit600_pct",          "Kit From 600 %",      "Custom-kit source: must total 100 with Kit From 400 %"),
    ("kit400_pct",          "Kit From 400 %",      "Custom-kit source: must total 100 with Kit From 600 %"),
]

HEADER_FILL  = PatternFill("solid", start_color="1F2937", end_color="1F2937")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
SUBLABEL_FONT= Font(color="6B7280", italic=True, size=9, name="Calibri")
TITLE_FONT   = Font(bold=True, size=16, name="Calibri", color="1F2937")
SECTION_FONT = Font(bold=True, size=12, name="Calibri", color="4F6EF7")
BODY_FONT    = Font(size=11, name="Calibri")
KEY_FILL     = PatternFill("solid", start_color="F3F4F6", end_color="F3F4F6")
THIN_BORDER  = Border(*[Side(style="thin", color="D1D5DB")] * 4)


def config_to_excel_bytes(areas, order_types) -> bytes:
    """Build a formatted .xlsx template: Instructions, Areas, Order Types sheets."""
    wb = Workbook()

    # ── Instructions sheet ───────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 95

    ws["A1"] = "WAREHOUSE CAPACITY PLANNER — Configuration Template"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Fill in the Areas and Order Types sheets, save the file, then upload it in the app's Settings page."
    ws["A2"].font = BODY_FONT
    r = 4
    for line in [
        "HOW THIS FILE IS ORGANIZED",
        "  •  Areas sheet        — physical dimensions for each storage area",
        "  •  Order Types sheet  — daily volume and percentage splits for each order type",
        "",
        "RULES",
        "  •  Do not rename Area ID or Order ID values — they are used as keys",
        "  •  Do not add or remove columns",
        "  •  Paper % + Consumable % must total 100 for every order type",
        "  •  Cust 1 % + Cust 2 % must total 100 for every order type",
        "  •  Direct to Packout % + Kitting % must total 100 for every order type",
        "  •  Kit 300 % + Kit 200 % must total 100 for every order type",
        "",
        "TIP",
        "  •  Download this template anytime from Settings to get your current live values pre-filled.",
    ]:
        ws.cell(row=r, column=1, value=line).font = (
            SECTION_FONT if line and line == line.upper() and not line.startswith(" ") else BODY_FONT
        )
        r += 1

    # ── Areas sheet ───────────────────────────────────────────────────────────
    ws_a = wb.create_sheet("Areas")
    ws_a.sheet_view.showGridLines = False
    for col_idx, (_, label, sub) in enumerate(AREA_COLS, start=1):
        c = ws_a.cell(row=1, column=col_idx, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER
        sub_c = ws_a.cell(row=2, column=col_idx, value=sub)
        sub_c.font = SUBLABEL_FONT
        sub_c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_a.column_dimensions[get_column_letter(col_idx)].width = 22
    ws_a.row_dimensions[1].height = 32
    ws_a.row_dimensions[2].height = 28
    ws_a.freeze_panes = "A3"

    for r_idx, a in enumerate(areas, start=3):
        values = [
            a.id, a.name, a.zone,
            # dimensions in feet (internal storage unit)
            round(a.rack_length_cuft, 2), round(a.rack_depth_cuft, 2),
            round(a.rack_height_cuft, 2), a.num_racks,
            round(a.box_length_cuft, 2), round(a.box_depth_cuft, 2),
            round(a.box_height_cuft, 2),
            a.efficiency, a.units_per_box,
            getattr(a, "of1_rack_length_cuft", 0), getattr(a, "of1_rack_depth_cuft", 0), getattr(a, "of1_rack_height_cuft", 0), getattr(a, "of1_num_racks", 0), getattr(a, "of2_rack_length_cuft", 0), getattr(a, "of2_rack_depth_cuft", 0), getattr(a, "of2_rack_height_cuft", 0), getattr(a, "of2_num_racks", 0),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws_a.cell(row=r_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if col_idx == 1:
                cell.fill = KEY_FILL

    # ── Order Types sheet ────────────────────────────────────────────────────
    ws_o = wb.create_sheet("Order Types")
    ws_o.sheet_view.showGridLines = False
    for col_idx, (_, label, sub) in enumerate(ORDER_COLS, start=1):
        c = ws_o.cell(row=1, column=col_idx, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER
        sub_c = ws_o.cell(row=2, column=col_idx, value=sub)
        sub_c.font = SUBLABEL_FONT
        sub_c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_o.column_dimensions[get_column_letter(col_idx)].width = 20
    ws_o.row_dimensions[1].height = 32
    ws_o.row_dimensions[2].height = 28
    ws_o.freeze_panes = "A3"

    for r_idx, ot in enumerate(order_types, start=3):
        values = [
            ot.id, ot.name, ot.daily_volume, ot.avg_units_per_order,
            ot.storage_split.paper_pct, ot.storage_split.consumable_pct,
            ot.customer_split.cust1_pct, ot.customer_split.cust2_pct,
            ot.kitting_split.packout_pct, ot.kitting_split.kitting_pct,
            ot.kit_storage_split.kit300_pct, ot.kit_storage_split.kit200_pct,
            ot.kit_source_split.kit600_pct, ot.kit_source_split.kit400_pct,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws_o.cell(row=r_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if col_idx == 1:
                cell.fill = KEY_FILL

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def excel_bytes_to_config(file_bytes):
    """Parse the Areas and Order Types sheets from an uploaded .xlsx template."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)

    if "Areas" not in wb.sheetnames:
        raise ValueError("Sheet 'Areas' not found. Make sure you uploaded the correct template file.")
    if "Order Types" not in wb.sheetnames:
        raise ValueError("Sheet 'Order Types' not found. Make sure you uploaded the correct template file.")

    ws_a = wb["Areas"]
    ws_o = wb["Order Types"]
    area_keys  = [k for k, _, _ in AREA_COLS]
    order_keys = [k for k, _, _ in ORDER_COLS]

    areas = []
    for row in ws_a.iter_rows(min_row=3, values_only=True):
        if row[0] is None or str(row[0]).strip() == "":
            continue
        d = dict(zip(area_keys, row))
        areas.append(StorageArea(
            id=str(d["area_id"]).strip(),
            name=str(d["area_name"]).strip(),
            zone=str(d["zone"]).strip(),
            # template columns are in feet (the internal unit)
            rack_length_cuft=float(d["rack_length_cuft"]),
            rack_depth_cuft=float(d["rack_depth_cuft"]),
            rack_height_cuft=float(d["rack_height_cuft"]),
            num_racks=int(float(d["num_racks"])),
            box_length_cuft=float(d["box_length_cuft"]),
            box_depth_cuft=float(d["box_depth_cuft"]),
            box_height_cuft=float(d["box_height_cuft"]),
            efficiency=float(d["efficiency"]),
            units_per_box=float(d["units_per_box"]),
            of1_rack_length_cuft=float(d.get("of1_rack_length_cuft", 0) or 0),
            of1_rack_depth_cuft=float(d.get("of1_rack_depth_cuft", 0) or 0),
            of1_rack_height_cuft=float(d.get("of1_rack_height_cuft", 0) or 0),
            of1_num_racks=int(float(d.get("of1_num_racks", 0) or 0)),
            of2_rack_length_cuft=float(d.get("of2_rack_length_cuft", 0) or 0),
            of2_rack_depth_cuft=float(d.get("of2_rack_depth_cuft", 0) or 0),
            of2_rack_height_cuft=float(d.get("of2_rack_height_cuft", 0) or 0),
            of2_num_racks=int(float(d.get("of2_num_racks", 0) or 0)),
        ))

    order_types = []
    for row in ws_o.iter_rows(min_row=3, values_only=True):
        if row[0] is None or str(row[0]).strip() == "":
            continue
        d = dict(zip(order_keys, row))
        order_types.append(OrderType(
            id=str(d["order_id"]).strip(),
            name=str(d["order_name"]).strip(),
            daily_volume=int(float(d["daily_volume"])),
            avg_units_per_order=int(float(d["avg_units_per_order"])),
            storage_split=StorageSplit(
                paper_pct=float(d["paper_pct"]),
                consumable_pct=float(d["consumable_pct"])),
            customer_split=CustomerSplit(
                cust1_pct=float(d["cust1_pct"]),
                cust2_pct=float(d["cust2_pct"])),
            kitting_split=KittingSplit(
                packout_pct=float(d["packout_pct"]),
                kitting_pct=float(d["kitting_pct"])),
            kit_storage_split=KitStorageSplit(
                kit300_pct=float(d.get("kit300_pct", 60.0)),
                kit200_pct=float(d.get("kit200_pct", 40.0))),
            kit_source_split=KitSourceSplit(
                kit600_pct=float(d.get("kit600_pct", 40.0)),
                kit400_pct=float(d.get("kit400_pct", 60.0))),
        ))

    if not areas:
        raise ValueError("No area rows found in the 'Areas' sheet (starting row 3).")
    if not order_types:
        raise ValueError("No order rows found in the 'Order Types' sheet (starting row 3).")

    return areas, order_types


if "areas" not in st.session_state:
    _loaded_areas, _loaded_orders = db.load_all()
    # Shared-analysis behavior: a brand-new or empty store seeds the built-in
    # defaults and saves them, so the first visitor sees a populated app and
    # every later visitor sees whatever was last saved.
    if not _loaded_areas and not _loaded_orders:
        _loaded_areas  = [copy.deepcopy(a) for a in DEFAULT_AREAS]
        _loaded_orders = [copy.deepcopy(o) for o in DEFAULT_ORDER_TYPES]
        try:
            db.save_all(_loaded_areas, _loaded_orders)
        except Exception:
            pass  # session-only if no DB configured; still populated for this user
    # Normalize dimensions to the grid's editing precision (2 dp for feet) so the
    # numbers shown in Settings and Analysis are always identical — no float drift.
    for _a in _loaded_areas:
        _a.rack_length_cuft = round(_a.rack_length_cuft, 2)
        _a.rack_depth_cuft  = round(_a.rack_depth_cuft, 2)
        _a.rack_height_cuft = round(_a.rack_height_cuft, 2)
        _a.box_length_cuft  = round(_a.box_length_cuft, 2)
        _a.box_depth_cuft   = round(_a.box_depth_cuft, 2)
        _a.box_height_cuft  = round(_a.box_height_cuft, 2)
    st.session_state.areas       = [copy.deepcopy(a) for a in _loaded_areas]
    st.session_state.order_types = [copy.deepcopy(o) for o in _loaded_orders]
    # Guarantee every order type has the kit-split fields, so order types loaded
    # or created before these fields existed can't crash any page on switch.
    for _o in st.session_state.order_types:
        if getattr(_o, "kit_storage_split", None) is None:
            _o.kit_storage_split = KitStorageSplit()
        if getattr(_o, "kit_source_split", None) is None:
            _o.kit_source_split = KitSourceSplit()
    # remember what's persisted so auto-save only fires on real edits
    st.session_state._saved_sig = _config_signature()

def get_engine():
    return WarehouseEngine(
        areas=[copy.deepcopy(a) for a in st.session_state.areas],
        order_types=[copy.deepcopy(o) for o in st.session_state.order_types],
    )


def analysis_report_excel_bytes(engine, multiplier: float) -> bytes:
    """Build a formatted, shareable Excel report of the current analysis."""
    snap = engine.snapshot(multiplier=multiplier)
    wb = Workbook()

    def _sheet_header(ws, headers, width=20):
        for ci, label in enumerate(headers, start=1):
            c = ws.cell(1, ci, label)
            c.font, c.fill = HEADER_FONT, HEADER_FILL
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"

    # ── Summary ───────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22
    ws["A1"] = "WAREHOUSE CAPACITY — ANALYSIS REPORT"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Generated " + datetime.now().strftime("%Y-%m-%d %H:%M")
                + "  ·  Volume multiplier x" + format(multiplier, ".1f"))
    ws["A2"].font = SUBLABEL_FONT
    summary = [
        ("Total capacity (boxes)",        snap.total_capacity_boxes),
        ("Total capacity (units)",        snap.total_capacity_units),
        ("Total load (boxes)",            int(snap.total_load_boxes)),
        ("Overall utilisation",           format(snap.overall_utilization, ".1f") + "%"),
        ("Areas over capacity",           len(snap.bottlenecks)),
        ("Areas near limit (85–100%)",    len(snap.warnings)),
    ]
    r = 4
    for label, val in summary:
        ws.cell(r, 1, label).font = BODY_FONT
        ws.cell(r, 2, val).font = Font(bold=True, size=11, name="Calibri")
        r += 1

    # ── Areas ─────────────────────────────────────────────────────────────────
    ws_a = wb.create_sheet("Areas")
    ws_a.sheet_view.showGridLines = False
    _sheet_header(ws_a, ["Area", "Zone", "Box L×W×H (ft)", "Capacity (boxes)",
                         "Load (boxes)", "Utilisation %", "Status",
                         "Capacity (units)", "Load (units)"])
    for ri, a in enumerate(snap.areas, start=2):
        vals = [
            a.area.name, a.area.zone,
            fmt_dims_ft(a.area.box_length_cuft, a.area.box_depth_cuft, a.area.box_height_cuft),
            a.capacity_boxes, int(a.load_boxes), round(a.utilization_pct, 1),
            status_label(a.utilization_pct), a.capacity_units, int(a.load_units),
        ]
        for ci, v in enumerate(vals, start=1):
            cell = ws_a.cell(ri, ci, v)
            cell.font, cell.border = BODY_FONT, THIN_BORDER

    # ── Zones ─────────────────────────────────────────────────────────────────
    zdf = engine.zone_summary(multiplier)
    ws_z = wb.create_sheet("Zones")
    ws_z.sheet_view.showGridLines = False
    zcols = ["zone_code", "zone_name", "areas", "capacity_boxes", "load_boxes",
             "capacity_units", "load_units", "utilization_pct"]
    zlabels = ["Zone", "Name", "Areas", "Capacity (boxes)", "Load (boxes)",
               "Capacity (units)", "Load (units)", "Utilisation %"]
    _sheet_header(ws_z, zlabels)
    for ri, (_, row) in enumerate(zdf.iterrows(), start=2):
        for ci, key in enumerate(zcols, start=1):
            cell = ws_z.cell(ri, ci, row[key])
            cell.font, cell.border = BODY_FONT, THIN_BORDER

    # ── Growth to capacity ────────────────────────────────────────────────────
    ws_b = wb.create_sheet("Growth to capacity")
    ws_b.sheet_view.showGridLines = False
    _sheet_header(ws_b, ["Area", "Reaches 85% at", "Reaches 100% at"], width=24)
    seq85  = {n: m for m, n, _ in engine.bottleneck_sequence(threshold_pct=85.0,  max_mult=50.0)}
    seq100 = {n: m for m, n, _ in engine.bottleneck_sequence(threshold_pct=100.0, max_mult=50.0)}
    names  = [a.area.name for a in snap.areas]
    for ri, name in enumerate(names, start=2):
        ws_b.cell(ri, 1, name).font = BODY_FONT
        ws_b.cell(ri, 2, ("x" + format(seq85[name], ".1f")) if name in seq85 else "—").font = BODY_FONT
        ws_b.cell(ri, 3, ("x" + format(seq100[name], ".1f")) if name in seq100 else "beyond x20").font = BODY_FONT
        for ci in range(1, 4):
            ws_b.cell(ri, ci).border = THIN_BORDER

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⬡ Warehouse Planner")
    st.markdown("---")
    page = st.radio("Navigate",
        ["Analysis", "Material flow", "Settings"],
        label_visibility="collapsed")
    st.markdown("---")
    if db.storage_mode() == "database":
        st.caption("🟢 Connected to database — changes persist across devices")
    else:
        st.caption("🟢 Saving locally — changes persist across refreshes on this server")
    st.markdown("---")
    st.markdown("**Quick status at x1.0**")
    if not st.session_state.get("areas"):
        st.info("No data yet — add it in Settings → Import / Export")
    else:
        _snap = get_engine().snapshot(1.0)
        nb, nw = len(_snap.bottlenecks), len(_snap.warnings)
        if nb:   st.error(f"⚠️ {nb} area(s) over capacity")
        elif nw: st.warning(f"△ {nw} area(s) near limit")
        else:    st.success("✓ All areas OK")
    st.markdown("---")
    st.caption("**Zone legend**")
    for zone, name in [("600","Paper"),("400","Consumables"),
                        ("300","Cust. Spec 1"),("200","Cust. Spec 2"),("100","Final")]:
        c = ZONE_COLORS.get(zone, "#888")
        st.markdown(
            '<span style="display:inline-block;width:10px;height:10px;'
            'border-radius:2px;background:' + c + ';margin-right:6px"></span>'
            '<small>' + zone + ' – ' + name + '</small>',
            unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
#  MATERIAL FLOW PAGE
# ═══════════════════════════════════════════════════════════════════════════════

if page == "Material flow":
    st.title("Material flow")
    render_last_updated()
    st.caption("How material moves through the warehouse — from inbound to final shipment.")

    if not st.session_state.get("areas"):
        st.info(
            "**No data yet.** Upload a data file in **Settings → Import / Export** to see the flow."
        )
        st.stop()

    def make_flow_diagram(engine):
        """Two clearly separated flows: the Standard order flow and the
        Custom-kit flow (added on top). Values are live at x1.0."""
        import math
        snap = engine.snapshot(1.0)
        fig  = go.Figure()

        # aggregate live quantities (boxes/units) at x1.0
        kit600 = sum(o.kit_units_from_paper()      for o in engine.order_types)
        kit400 = sum(o.kit_units_from_consumable() for o in engine.order_types)
        kit300 = sum(o.kit_units_to_300()          for o in engine.order_types)
        kit200 = sum(o.kit_units_to_200()          for o in engine.order_types)
        kittot = sum(o.kit_units_total()           for o in engine.order_types)

        def add_arrow(x0, y0, x1, y1, color, dash="solid", width=2.5):
            fig.add_trace(go.Scatter(x=[x0, x1], y=[y0, y1], mode="lines",
                line=dict(color=color, width=width, dash=dash),
                hoverinfo="skip", showlegend=False))
            angle = math.degrees(math.atan2(y1-y0, x1-x0))
            fig.add_trace(go.Scatter(x=[x1], y=[y1], mode="markers",
                marker=dict(symbol="arrow", size=15, color=color, angle=angle-90, line=dict(width=0)),
                hoverinfo="skip", showlegend=False))

        def add_node(x, y, title, subtitle, tc, bg, bc, w=20, h=8):
            fig.add_shape(type="rect", x0=x-w/2+0.5, y0=y-h/2-0.6, x1=x+w/2+0.5, y1=y+h/2-0.6,
                fillcolor="rgba(0,0,0,0.35)", line=dict(width=0))
            fig.add_shape(type="rect", x0=x-w/2, y0=y-h/2, x1=x+w/2, y1=y+h/2,
                fillcolor=bg, line=dict(color=bc, width=2))
            fig.add_annotation(x=x, y=y+(1.4 if subtitle else 0), text="<b>"+title+"</b>",
                showarrow=False, font=dict(size=14, color=tc), xref="x", yref="y")
            if subtitle:
                fig.add_annotation(x=x, y=y-2.0, text=subtitle,
                    showarrow=False, font=dict(size=10, color="#9ca3af"), xref="x", yref="y")

        def edge_label(x, y, text, color, size=11):
            fig.add_annotation(x=x, y=y, text="<i>"+text+"</i>", showarrow=False,
                font=dict(size=size, color=color), xref="x", yref="y",
                bgcolor="rgba(15,17,23,0.8)", borderpad=2)

        def column_header(x, text, color):
            fig.add_annotation(x=x, y=99, text="<b>"+text+"</b>", showarrow=False,
                font=dict(size=15, color=color), xref="x", yref="y",
                bgcolor="rgba(15,17,23,0.85)", bordercolor=color, borderwidth=1, borderpad=6)

        # dashed divider between the two flows
        fig.add_shape(type="line", x0=50, y0=2, x1=50, y1=94,
            line=dict(color="#2e3250", width=1, dash="dot"))

        # ══ LEFT: STANDARD FLOW ══════════════════════════════════════════
        LB = "#4f6ef7"
        column_header(25, "STANDARD ORDER FLOW", "#818cf8")
        add_node(25, 88, "Inbound Orders", "", "#c7d2fe", "#1e2235", "#4f6ef7", w=26, h=7)
        add_node(12, 68, "600 Paper", "raw paper", "#a0a8f0", "#1a2040", "#4f6ef7", w=18, h=8)
        add_node(38, 68, "400 Consumables", "raw consumables", "#67d8f0", "#0a2025", "#06b6d4", w=18, h=8)
        add_node(25, 44, "300 / 200", "customer areas", "#6ee7b7", "#0a2018", "#10b981", w=26, h=8)
        add_node(25, 16, "100 Final", "ready to ship", "#fca5a5", "#2a0a0a", "#ef4444", w=26, h=7)
        add_arrow(20, 84.5, 12, 72.5, LB)
        add_arrow(30, 84.5, 38, 72.5, "#06b6d4")
        add_arrow(12, 63.5, 22, 48.5, LB)
        add_arrow(38, 63.5, 28, 48.5, "#06b6d4")
        add_arrow(25, 39.5, 25, 19.5, "#10b981")
        edge_label(13, 78, "Paper %", LB)
        edge_label(37, 78, "Consumable %", "#06b6d4")
        edge_label(25, 55, "Customer split (300/200)", "#10b981")
        edge_label(25, 28, "packout", "#ef4444")

        # ══ RIGHT: CUSTOM-KIT FLOW ═══════════════════════════════════════
        KC = "#f59e0b"
        column_header(75, "CUSTOM-KIT FLOW", "#fbbf24")
        add_node(75, 88, "Kitting % of orders", "custom kits", "#fde68a", "#2a1800", "#f59e0b", w=28, h=7)
        add_node(62, 68, "600 Paper", "kit material", "#a0a8f0", "#1a2040", "#4f6ef7", w=18, h=8)
        add_node(88, 68, "400 Consumables", "kit material", "#67d8f0", "#0a2025", "#06b6d4", w=18, h=8)
        add_node(75, 50, "Kitting", "custom kit assembly", "#d1d5db", "#111827", "#9ca3af", w=24, h=7)
        add_node(75, 32, "300 / 200", "kitted storage", "#6ee7b7", "#0a2018", "#10b981", w=26, h=7)
        add_node(75, 12, "100 Final", "ready to ship", "#fca5a5", "#2a0a0a", "#ef4444", w=26, h=7)
        add_arrow(70, 84.5, 62, 72.5, KC)
        add_arrow(80, 84.5, 88, 72.5, KC)
        add_arrow(62, 63.5, 72, 53.5, KC)
        add_arrow(88, 63.5, 78, 53.5, KC)
        add_arrow(75, 46.5, 75, 35.5, KC)
        add_arrow(75, 28.5, 75, 15.5, KC)
        edge_label(60, 78, "pull " + str(int(kit600)) + "u", KC)
        edge_label(90, 78, "pull " + str(int(kit400)) + "u", KC)
        edge_label(75, 59, "by storage split", KC)
        edge_label(75, 41, "kit split (" + str(int(kit300)) + "/" + str(int(kit200)) + "u)", "#10b981")
        edge_label(75, 22, str(int(kittot)) + "u to ship", "#ef4444")

        fig.update_layout(
            height=560, margin=dict(l=6, r=6, t=30, b=6),
            paper_bgcolor="#0f1117", plot_bgcolor="#0f1117",
            xaxis=dict(visible=False, range=[-2, 102], fixedrange=True),
            yaxis=dict(visible=False, range=[0, 104], fixedrange=True),
            showlegend=False)
        return fig

    st.plotly_chart(make_flow_diagram(get_engine()), width="stretch")
    st.markdown("---")

    engine = get_engine()
    snap   = engine.snapshot(1.0)
    fl     = snap.flow

    st.subheader("Live flow volumes at x1.0")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Paper to Zone 300",       str(int(fl.paper_to_300))         + " boxes/day")
    c2.metric("Paper to Zone 200",       str(int(fl.paper_to_200))         + " boxes/day")
    c3.metric("Consumables to Zn 300",   str(int(fl.consumables_to_300))   + " boxes/day")
    c4.metric("Consumables to Zn 200",   str(int(fl.consumables_to_200))   + " boxes/day")

    st.markdown("#### Per order type")
    st.caption("Expand any order type to see its splits.")
    for ot in engine.order_types:
        with st.expander("**" + ot.name + "**", expanded=False):
            r1, r2, r3 = st.columns(3)
            r1.metric("Daily orders",    ot.daily_volume)
            r2.metric("Avg units/order", ot.avg_units_per_order)
            r3.metric("Total units/day", str(int(ot.total_units())))

            st.caption("Split 1 - Storage")
            s1a, s1b = st.columns(2)
            s1a.metric("600 Paper %",        str(int(ot.storage_split.paper_pct))      + "%")
            s1b.metric("400 Consumables %",  str(int(ot.storage_split.consumable_pct)) + "%")

            st.caption("Split 2 - Customer (of the Consumables portion)")
            s2a, s2b = st.columns(2)
            s2a.metric("to Zone 300 %", str(int(ot.customer_split.cust1_pct)) + "%")
            s2b.metric("to Zone 200 %", str(int(ot.customer_split.cust2_pct)) + "%")

            st.caption("Split 3 - Kitting (of 300/200 material)")
            s3a, s3b, s3c = st.columns(3)
            s3a.metric("Direct to packout %", str(int(ot.kitting_split.packout_pct)) + "%")
            s3b.metric("to Kitting %",        str(int(ot.kitting_split.kitting_pct)) + "%")
            kit_boxes = sum(
                ot.boxes_in_area(a, 1.0) * (ot.kitting_split.kitting_pct / 100)
                for a in engine.areas if a.zone in ("300", "200")
            )
            s3c.metric("Kitting boxes/day", str(round(kit_boxes, 1)))

    with st.expander("📋 Flow rules reference", expanded=False):
        st.markdown("""
| Rule | Logic |
|---|---|
| **Split 1 - Storage** | 600% + 400% = 100% — where the order draws its units from |
| **Rule 2 - Paper path** | 600 to Smart Bulk (same volume staged) then direct to 100 Final — bypasses 300/200 |
| **Split 2 - Customer** | 300% + 200% = 100% — how the 400 consumable portion divides between customers |
| **Split 3 - Kitting** | Packout% + Kitting% = 100% — how 300/200 material routes onward |
| **Rule 4 - Kitting loop** | Kitting returns to same zone (300 or 200) then flows to 100 normally |
| **Rule 5 - Final** | All paths converge at 100 Final/Packout before shipping |
        """)


# ═══════════════════════════════════════════════════════════════════════════════
#  SETTINGS PAGE  — section renderers (one per sub-tab)
# ═══════════════════════════════════════════════════════════════════════════════

def _render_unit_converter():
    st.caption("Enter a value in any unit to see all equivalents instantly.")
    cc1, cc2 = st.columns(2)
    conv_val = cc1.number_input("Value", value=1.0, step=0.1, key="conv_val", min_value=0.0)
    conv_from = cc2.selectbox("From unit", list(UNITS.keys()), key="conv_from")
    val_in_cuft = conv_val * UNITS[conv_from]["to_cuft"]
    st.markdown("**Equivalents:**")
    rc = st.columns(4)
    for i, (ukey, udata) in enumerate(UNITS.items()):
        converted = val_in_cuft * udata["from_cuft"]
        if ukey == conv_from:
            rc[i].metric(udata["label"], f"{converted:,.4f}", delta="← input", delta_color="off")
        else:
            rc[i].metric(udata["label"], f"{converted:,.4f}")


def _render_area_settings():
    """Editable grid of ALL storage areas — one row each, no vertical scrolling."""
    st.subheader("Storage areas")
    if not st.session_state.areas:
        st.info(
            "No storage areas yet. Upload a data file in **Import / Export**, then edit it here."
        )
        return {}
    st.caption(
        "Every area is one row — edit any cell directly. Rack and box dimensions are in "
        "**feet (ft)**; volume shows in **ft³**. Capacity updates live in the preview "
        "below; changes save automatically."
    )

    df = pd.DataFrame([{
        "ID":        a.id,
        "Area":      a.name,
        "Zone":      a.zone,
        "Rack L":    round(a.rack_length_cuft, 2),
        "Rack D":    round(a.rack_depth_cuft, 2),
        "Rack H":    round(a.rack_height_cuft, 2),
        "# Racks":   int(a.num_racks),
        "Box L":     round(a.box_length_cuft, 2),
        "Box W":     round(a.box_depth_cuft, 2),
        "Box H":     round(a.box_height_cuft, 2),
        "Eff.":      float(a.efficiency),
        "Units/box": float(a.units_per_box),
    } for a in st.session_state.areas]).set_index("ID")

    row_h    = 29
    grid_h   = int((len(df) + 1) * row_h + 3)
    edited = st.data_editor(
        df, key="areas_editor", hide_index=True, num_rows="fixed",
        width="stretch", height=grid_h, row_height=row_h,
        column_config={
            "Area":      st.column_config.TextColumn("Area", width="medium", help="Display name for this storage area."),
            "Zone":      st.column_config.TextColumn("Zone", disabled=True, width="small", help="Flow zone (fixed)."),
            "Rack L":    st.column_config.NumberColumn("Rack L (ft)", min_value=0.01, step=0.5, format="%.2f", help="Length of one rack, in feet."),
            "Rack D":    st.column_config.NumberColumn("Rack D (ft)", min_value=0.01, step=0.5, format="%.2f", help="Depth of one rack, in feet."),
            "Rack H":    st.column_config.NumberColumn("Rack H (ft)", min_value=0.01, step=0.5, format="%.2f", help="Height of one rack, in feet."),
            "# Racks":   st.column_config.NumberColumn("# Racks",  min_value=1,     step=1,    format="%d", help="How many racks of this size are in the area."),
            "Box L":     st.column_config.NumberColumn("Box L (ft)", min_value=0.01, step=0.1, format="%.2f", help="Length of the typical box/pallet stored here."),
            "Box W":     st.column_config.NumberColumn("Box W (ft)", min_value=0.01, step=0.1, format="%.2f", help="Width of the typical box/pallet."),
            "Box H":     st.column_config.NumberColumn("Box H (ft)", min_value=0.01, step=0.1, format="%.2f", help="Height of the typical box/pallet."),
            "Eff.":      st.column_config.NumberColumn("Eff. (0–1)", min_value=0.1, max_value=1.0, step=0.01, format="%.2f", help="Usable fraction of rack volume after aisles/reach space. 0.75 = 75%."),
            "Units/box": st.column_config.NumberColumn("Units/box", min_value=1.0, step=1.0, format="%.0f", help="Average number of individual units in one box."),
        },
    )

    # ── build updates + live capacity preview from the edited grid ────────────
    # Dimensions are already in feet (the internal unit).
    area_updates, prev_rows = {}, []
    for aid, row in edited.iterrows():
        rl, rd, rh = float(row["Rack L"]), float(row["Rack D"]), float(row["Rack H"])
        bl, bw, bh = float(row["Box L"]),  float(row["Box W"]),  float(row["Box H"])
        racks = int(row["# Racks"]); eff = float(row["Eff."]); upb = float(row["Units/box"])

        total_vol = rl * rd * rh * racks
        box_vol   = bl * bw * bh
        cap       = int((total_vol * eff) / box_vol) if box_vol > 0 else 0

        area_updates[aid] = dict(
            name=str(row["Area"]),
            rack_length_cuft=rl, rack_depth_cuft=rd, rack_height_cuft=rh,
            num_racks=racks,
            box_length_cuft=bl, box_depth_cuft=bw, box_height_cuft=bh,
            efficiency=eff, units_per_box=upb,
        )
        prev_rows.append({
            "Area":                 str(row["Area"]),
            "Area volume (ft³)":    round(total_vol, 1),
            "Capacity (boxes)":     cap,
            "Capacity (units)":     int(cap * upb),
        })

    st.caption("**Live capacity preview** — recalculates as you type, applied on Save:")
    prev_df = pd.DataFrame(prev_rows)
    st.dataframe(prev_df, hide_index=True, width="stretch",
                 height=int((len(prev_df) + 1) * row_h + 3), row_height=row_h)

    # ── Overflow zones (2 per area) ──────────────────────────────────────────
    st.markdown("---")
    st.markdown("#### Overflow zones")
    st.caption(
        "Each area can have up to two overflow zones. When the main area fills, load "
        "spills into overflow 1, then overflow 2. Overflow capacity uses the same box "
        "size and efficiency as the main area. Leave racks at 0 for no overflow. "
        "An area shows **WARNING** once it spills into overflow, and **OVER CAPACITY** "
        "only when the main area plus both overflow zones are full."
    )
    area_map = {a.id: a for a in st.session_state.areas}
    for aid in area_updates:
        a = area_map.get(aid)
        if a is None:
            continue
        u = area_updates[aid]
        box_vol = u["box_length_cuft"] * u["box_depth_cuft"] * u["box_height_cuft"]
        eff = u["efficiency"]
        def _cap(rl, rd, rh, nr):
            v = rl * rd * rh * nr
            return int((v * eff) / box_vol) if box_vol > 0 else 0
        main_cap = u_cap = _cap(u["rack_length_cuft"], u["rack_depth_cuft"], u["rack_height_cuft"], u["num_racks"])
        with st.expander(a.name + "  —  overflow zones", expanded=False):
            for n in (1, 2):
                st.markdown("**Overflow " + str(n) + "**")
                c1, c2, c3, c4, c5 = st.columns([1, 1, 1, 1, 1.4])
                rl = c1.number_input("Rack L (ft)", min_value=0.0, step=0.5, format="%.2f",
                    value=float(getattr(a, f"of{n}_rack_length_cuft", 0.0)), key=f"of{n}_rl_{aid}")
                rd = c2.number_input("Rack D (ft)", min_value=0.0, step=0.5, format="%.2f",
                    value=float(getattr(a, f"of{n}_rack_depth_cuft", 0.0)), key=f"of{n}_rd_{aid}")
                rh = c3.number_input("Rack H (ft)", min_value=0.0, step=0.5, format="%.2f",
                    value=float(getattr(a, f"of{n}_rack_height_cuft", 0.0)), key=f"of{n}_rh_{aid}")
                nr = c4.number_input("# Racks", min_value=0, step=1, format="%d",
                    value=int(getattr(a, f"of{n}_num_racks", 0)), key=f"of{n}_nr_{aid}")
                ofcap = _cap(rl, rd, rh, nr)
                c5.metric("Overflow " + str(n) + " cap", f"{ofcap:,} boxes")
                u[f"of{n}_rack_length_cuft"] = rl
                u[f"of{n}_rack_depth_cuft"]  = rd
                u[f"of{n}_rack_height_cuft"] = rh
                u[f"of{n}_num_racks"]        = int(nr)
            of1 = _cap(u["of1_rack_length_cuft"], u["of1_rack_depth_cuft"], u["of1_rack_height_cuft"], u["of1_num_racks"])
            of2 = _cap(u["of2_rack_length_cuft"], u["of2_rack_depth_cuft"], u["of2_rack_height_cuft"], u["of2_num_racks"])
            st.caption(
                "Total effective capacity: **" + f"{main_cap + of1 + of2:,}" + " boxes** "
                "(main " + f"{main_cap:,}" + " + overflow " + f"{of1 + of2:,}" + ")"
            )
    return area_updates


def _render_order_settings():
    """Editable grid of ALL order types — one row each, with live split checks."""
    st.subheader("Order types")
    if not st.session_state.order_types:
        st.info(
            "No order types yet. Upload a data file in **Import / Export**, then edit it here."
        )
        return {}, True
    st.caption(
        "Every order type is one row — edit any cell directly. Both split pairs must "
        "each total 100% (600+400, 300+200). Kitting is configured in the Kitting tab. "
        "Changes apply automatically."
    )

    df = pd.DataFrame([{
        "ID":        ot.id,
        "Order":     ot.name,
        "Daily vol": int(ot.daily_volume),
        "Avg units": int(ot.avg_units_per_order),
        "600 %":     float(ot.storage_split.paper_pct),
        "400 %":     float(ot.storage_split.consumable_pct),
        "300 %":     float(ot.customer_split.cust1_pct),
        "200 %":     float(ot.customer_split.cust2_pct),
    } for ot in st.session_state.order_types]).set_index("ID")

    def pct(label, help=None):
        return st.column_config.NumberColumn(label, min_value=0.0, max_value=100.0, step=1.0, format="%.0f", help=help)

    row_h  = 29
    grid_h = int((len(df) + 1) * row_h + 3)
    edited = st.data_editor(
        df, key="orders_editor", hide_index=True, num_rows="fixed",
        width="stretch", height=grid_h, row_height=row_h,
        column_config={
            "Order":     st.column_config.TextColumn("Order", width="medium", help="Order type name/code."),
            "Daily vol": st.column_config.NumberColumn("Daily vol", min_value=1, step=1, format="%d", help="Number of orders of this type per day."),
            "Avg units": st.column_config.NumberColumn("Avg units", min_value=1, step=1, format="%d", help="Average units per order."),
            "600 %":     pct("600 Paper %",  "Share of units drawn from Paper (600). 600% + 400% must total 100%."),
            "400 %":     pct("400 Cons %",   "Share drawn from Consumables (400). 600% + 400% must total 100%."),
            "300 %":     pct("300 Cust1 %",  "Of the consumables, share to Customer zone 300. 300% + 200% must total 100%."),
            "200 %":     pct("200 Cust2 %",  "Of the consumables, share to Customer zone 200. 300% + 200% must total 100%."),
        },
    )

    # ── build updates + live split validation from the edited grid ────────────
    order_updates, check_rows, all_ok = {}, [], True
    for oid, row in edited.iterrows():
        s1 = float(row["600 %"]) + float(row["400 %"])
        s2 = float(row["300 %"]) + float(row["200 %"])
        ok = all(abs(s - 100) <= 0.5 for s in (s1, s2))
        all_ok = all_ok and ok
        check_rows.append({
            "Order":           str(row["Order"]),
            "Storage 600+400": f"{s1:.0f}%",
            "Cust 300+200":    f"{s2:.0f}%",
            "Units/day":       int(row["Daily vol"]) * int(row["Avg units"]),
            "Valid":           "✅" if ok else "⚠️",
        })
        order_updates[oid] = dict(
            name=str(row["Order"]),
            daily_volume=int(row["Daily vol"]),
            avg_units_per_order=int(row["Avg units"]),
            paper_pct=float(row["600 %"]),   consumable_pct=float(row["400 %"]),
            cust1_pct=float(row["300 %"]),   cust2_pct=float(row["200 %"]),
        )

    st.caption("**Split validation** — each pair must total 100%:")
    check_df = pd.DataFrame(check_rows)
    st.dataframe(check_df, hide_index=True, width="stretch",
                 height=int((len(check_df) + 1) * row_h + 3), row_height=row_h)
    if all_ok:
        st.success("All split pairs total 100%.")
    else:
        st.warning("Some split pairs don't total 100% (see ⚠️ rows) — auto-save is paused until every pair totals 100%.")

    # ── Add a new order type ─────────────────────────────────────────────────
    st.markdown("---")
    with st.expander("➕  Add a new order type", expanded=False):
        st.caption(
            "Give the order type a short code and name, its daily volume, and how it "
            "splits across the zones. Each split pair must total 100%. "
            "You can also add order types in bulk via Import / Export."
        )
        existing_ids = {o.id.upper() for o in st.session_state.order_types}

        c1, c2, c3 = st.columns([1, 2, 1])
        new_id   = c1.text_input("Code", key="new_ot_id", placeholder="e.g. CX", max_chars=6)
        new_name = c2.text_input("Name", key="new_ot_name", placeholder="e.g. CX – Customer service")
        new_vol  = c3.number_input("Daily vol", min_value=1, value=10, step=1, key="new_ot_vol")

        c4, c5 = st.columns(2)
        new_units = c4.number_input("Avg units / order", min_value=1, value=10, step=1, key="new_ot_units")

        st.caption("Storage split (600 + 400 = 100)")
        s1a, s1b = st.columns(2)
        p_paper = s1a.number_input("600 Paper %", 0.0, 100.0, 40.0, 1.0, key="new_ot_paper")
        p_cons  = s1b.number_input("400 Consumable %", 0.0, 100.0, 60.0, 1.0, key="new_ot_cons")

        st.caption("Customer split (300 + 200 = 100)")
        s2a, s2b = st.columns(2)
        p_c1 = s2a.number_input("300 Cust 1 %", 0.0, 100.0, 60.0, 1.0, key="new_ot_c1")
        p_c2 = s2b.number_input("200 Cust 2 %", 0.0, 100.0, 40.0, 1.0, key="new_ot_c2")

        st.caption("Kitting split (Packout + Kitting = 100)")
        s3a, s3b = st.columns(2)
        p_pack = s3a.number_input("Packout %", 0.0, 100.0, 80.0, 1.0, key="new_ot_pack")
        p_kit  = s3b.number_input("Kitting %", 0.0, 100.0, 20.0, 1.0, key="new_ot_kit")

        st.caption("Custom-kit storage split (Kit 300 + Kit 200 = 100)")
        s4a, s4b = st.columns(2)
        p_k300 = s4a.number_input("Kit 300 %", 0.0, 100.0, 60.0, 1.0, key="new_ot_k300")
        p_k200 = s4b.number_input("Kit 200 %", 0.0, 100.0, 40.0, 1.0, key="new_ot_k200")

        # validation
        problems = []
        code = (new_id or "").strip().upper()
        if not code:
            problems.append("Enter a code.")
        elif code in existing_ids:
            problems.append("Code '" + code + "' is already used — pick a unique one.")
        if abs((p_paper + p_cons) - 100) > 0.5:
            problems.append("Storage split (600 + 400) must total 100.")
        if abs((p_c1 + p_c2) - 100) > 0.5:
            problems.append("Customer split (300 + 200) must total 100.")
        if abs((p_pack + p_kit) - 100) > 0.5:
            problems.append("Kitting split (Packout + Kitting) must total 100.")
        if abs((p_k300 + p_k200) - 100) > 0.5:
            problems.append("Custom-kit storage split (Kit 300 + Kit 200) must total 100.")

        if problems:
            for p in problems:
                st.caption("• " + p)

        if st.button("Add order type", type="primary", disabled=bool(problems), key="add_ot_btn"):
            st.session_state.order_types.append(OrderType(
                id=code,
                name=(new_name.strip() or code),
                daily_volume=int(new_vol),
                avg_units_per_order=int(new_units),
                storage_split=StorageSplit(paper_pct=float(p_paper), consumable_pct=float(p_cons)),
                customer_split=CustomerSplit(cust1_pct=float(p_c1), cust2_pct=float(p_c2)),
                kitting_split=KittingSplit(packout_pct=float(p_pack), kitting_pct=float(p_kit)),
                kit_storage_split=KitStorageSplit(kit300_pct=float(p_k300), kit200_pct=float(p_k200)),
                kit_source_split=KitSourceSplit(kit600_pct=40.0, kit400_pct=60.0),
            ))
            # persist immediately and refresh the saved fingerprint
            st.session_state._saved_sig = _config_signature()
            db.save_all(st.session_state.areas, st.session_state.order_types)
            # clear the input fields for the next entry
            for k in ("new_ot_id", "new_ot_name"):
                st.session_state.pop(k, None)
            st.success("Added order type '" + code + "'.")
            st.rerun()

    return order_updates, all_ok


def _render_kitting_settings():
    """Separate table for the custom-kit loop, per order type:
    the kit % of orders, where kit material is pulled from (600/400),
    and where kitted material is stored (300/200)."""
    st.subheader("Kitting — custom kit loop")
    if not st.session_state.order_types:
        st.info("No order types yet. Add them in the Order Types tab first.")
        return {}, True
    st.caption(
        "For each order type: the share of orders that become custom kits, where the "
        "kit material is pulled from (Paper 600 / Consumables 400), and where the kitted "
        "material is stored (300 / 200). Kit 600+400 and Kit 300+200 must each total 100%. "
        "Kitted material then ships to 100."
    )

    df = pd.DataFrame([{
        "ID":        ot.id,
        "Order":     ot.name,
        "Kit %":     float(ot.kitting_split.kitting_pct),
        "From 600 %": float(ot.kit_source_split.kit600_pct),
        "From 400 %": float(ot.kit_source_split.kit400_pct),
        "To 300 %":  float(ot.kit_storage_split.kit300_pct),
        "To 200 %":  float(ot.kit_storage_split.kit200_pct),
    } for ot in st.session_state.order_types]).set_index("ID")

    def pct(label, help=None):
        return st.column_config.NumberColumn(label, min_value=0.0, max_value=100.0, step=1.0, format="%.0f", help=help)

    row_h  = 29
    grid_h = int((len(df) + 1) * row_h + 3)
    edited = st.data_editor(
        df, key="kitting_editor", hide_index=True, num_rows="fixed",
        width="stretch", height=grid_h, row_height=row_h,
        column_config={
            "Order":      st.column_config.TextColumn("Order", width="medium", disabled=True),
            "Kit %":      pct("Kit % of orders", "Share of this order type's volume that goes through the custom-kit loop."),
            "From 600 %": pct("From 600 %", "Kit material pulled from Paper (600). From 600% + From 400% must total 100%."),
            "From 400 %": pct("From 400 %", "Kit material pulled from Consumables (400). From 600% + From 400% must total 100%."),
            "To 300 %":   pct("To 300 %",   "Kitted material stored in 300. To 300% + To 200% must total 100%."),
            "To 200 %":   pct("To 200 %",   "Kitted material stored in 200. To 300% + To 200% must total 100%."),
        },
    )

    kit_updates, check_rows, all_ok = {}, [], True
    for oid, row in edited.iterrows():
        src = float(row["From 600 %"]) + float(row["From 400 %"])
        sto = float(row["To 300 %"]) + float(row["To 200 %"])
        ok = all(abs(s - 100) <= 0.5 for s in (src, sto))
        all_ok = all_ok and ok
        check_rows.append({
            "Order":        str(row["Order"]),
            "Kit %":        f"{float(row['Kit %']):.0f}%",
            "Source 600+400": f"{src:.0f}%",
            "Store 300+200":  f"{sto:.0f}%",
            "Valid":        "✅" if ok else "⚠️",
        })
        kit_updates[oid] = dict(
            kitting_pct=float(row["Kit %"]),
            packout_pct=100.0 - float(row["Kit %"]),   # complement keeps Split 3 consistent
            kit600_pct=float(row["From 600 %"]), kit400_pct=float(row["From 400 %"]),
            kit300_pct=float(row["To 300 %"]),   kit200_pct=float(row["To 200 %"]),
        )

    st.caption("**Kit split validation** — each pair must total 100%:")
    check_df = pd.DataFrame(check_rows)
    st.dataframe(check_df, hide_index=True, width="stretch",
                 height=int((len(check_df) + 1) * row_h + 3), row_height=row_h)
    if all_ok:
        st.success("All kit split pairs total 100%.")
    else:
        st.warning("Some kit split pairs don't total 100% (see ⚠️ rows) — auto-save is paused until they do.")

    return kit_updates, all_ok


def _apply_updates(area_updates, order_updates):
    """Write the collected edits back into the in-memory session state only."""
    area_map = {a.id: a for a in st.session_state.areas}
    for aid, u in area_updates.items():
        a = area_map.get(aid)
        if a is None:
            continue
        a.name                = u.get("name", a.name)
        a.rack_length_cuft    = u["rack_length_cuft"]
        a.rack_depth_cuft     = u["rack_depth_cuft"]
        a.rack_height_cuft    = u["rack_height_cuft"]
        a.num_racks           = u["num_racks"]
        a.box_length_cuft     = u["box_length_cuft"]
        a.box_depth_cuft      = u["box_depth_cuft"]
        a.box_height_cuft     = u["box_height_cuft"]
        a.efficiency          = u["efficiency"]
        a.units_per_box       = u["units_per_box"]
        for _k in ("of1_rack_length_cuft","of1_rack_depth_cuft","of1_rack_height_cuft","of1_num_racks",
                   "of2_rack_length_cuft","of2_rack_depth_cuft","of2_rack_height_cuft","of2_num_racks"):
            if _k in u:
                setattr(a, _k, u[_k])

    ot_map = {o.id: o for o in st.session_state.order_types}
    for oid, u in order_updates.items():
        ot = ot_map.get(oid)
        if ot is None:
            continue
        ot.name                = u.get("name", ot.name)
        ot.daily_volume        = u["daily_volume"]
        ot.avg_units_per_order = u["avg_units_per_order"]
        ot.storage_split  = StorageSplit(paper_pct=u["paper_pct"],      consumable_pct=u["consumable_pct"])
        ot.customer_split = CustomerSplit(cust1_pct=u["cust1_pct"],     cust2_pct=u["cust2_pct"])
        # Kitting fields now come from the separate Kitting table (merged into
        # this dict). Fall back to the order type's current values if absent.
        ot.kitting_split = KittingSplit(
            packout_pct=u.get("packout_pct", ot.kitting_split.packout_pct),
            kitting_pct=u.get("kitting_pct", ot.kitting_split.kitting_pct))
        ot.kit_storage_split = KitStorageSplit(
            kit300_pct=u.get("kit300_pct", ot.kit_storage_split.kit300_pct),
            kit200_pct=u.get("kit200_pct", ot.kit_storage_split.kit200_pct))
        ot.kit_source_split = KitSourceSplit(
            kit600_pct=u.get("kit600_pct", ot.kit_source_split.kit600_pct),
            kit400_pct=u.get("kit400_pct", ot.kit_source_split.kit400_pct))


def _autosave_if_changed(valid: bool):
    """Persist automatically when the config changed and is valid.

    Returns a status string: 'saved', 'unchanged', or 'blocked'.
    """
    sig = _config_signature()
    if sig == st.session_state.get("_saved_sig"):
        return "unchanged"
    if not valid:
        return "blocked"
    if db.save_all(st.session_state.areas, st.session_state.order_types):
        st.session_state._saved_sig = sig
        return "saved"
    return "blocked"


def _render_excel_io():
    st.subheader("📄 Excel setup file")
    st.caption(
        "Download the current setup as an Excel file, edit it if you like, then "
        "upload it here to load and share it. This is how setups move in and out "
        "of the app."
    )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    excel_data = config_to_excel_bytes(st.session_state.areas, st.session_state.order_types)

    col_dl, col_ul = st.columns(2)

    with col_dl:
        st.markdown("**Step 1 — Download template**")
        st.caption(
            "Contains your current values pre-filled across three sheets: "
            "Instructions, Areas, and Order Types. Edit numbers directly, save, and upload."
        )
        st.download_button(
            "⬇️ Download configuration template (.xlsx)",
            data=excel_data,
            file_name="warehouse_config_" + timestamp + ".xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )

    with col_ul:
        st.markdown("**Step 2 — Upload filled template**")
        st.caption("Upload the same .xlsx after editing. Both sheets load at once.")
        uploaded = st.file_uploader(
            "Choose Excel file", type=["xlsx"], key="upload_config",
            label_visibility="collapsed")

        if uploaded is not None:
            try:
                preview_areas, preview_orders = excel_bytes_to_config(uploaded.getvalue())
                st.success(
                    "✅ File parsed successfully — " + str(len(preview_areas)) + " areas and "
                    + str(len(preview_orders)) + " order types found. Review below, then click Apply.")

                with st.expander("📋 Preview imported values", expanded=True):
                    st.markdown("**Areas**")
                    st.dataframe(
                        pd.DataFrame([{
                            "Area": a.name, "Zone": a.zone,
                            "Rack L×D×H (ft)": fmt_dims_ft(a.rack_length_cuft, a.rack_depth_cuft, a.rack_height_cuft),
                            "Racks": a.num_racks,
                            "Volume (ft³)": round(a.volume_cuft, 1),
                            "Box L×W×H (ft)": fmt_dims_ft(a.box_length_cuft, a.box_depth_cuft, a.box_height_cuft),
                            "Efficiency": a.efficiency,
                            "Units/box": a.units_per_box,
                        } for a in preview_areas]),
                        width="stretch", hide_index=True,
                        height=int((len(preview_areas) + 1) * 29 + 3), row_height=29)
                    st.markdown("**Order types**")
                    st.dataframe(
                        pd.DataFrame([{
                            "Order": ot.name,
                            "Daily vol": ot.daily_volume,
                            "Avg units": ot.avg_units_per_order,
                            "Paper %": ot.storage_split.paper_pct,
                            "Consumable %": ot.storage_split.consumable_pct,
                            "Cust1 %": ot.customer_split.cust1_pct,
                            "Cust2 %": ot.customer_split.cust2_pct,
                            "Packout %": ot.kitting_split.packout_pct,
                            "Kitting %": ot.kitting_split.kitting_pct,
                        } for ot in preview_orders]),
                        width="stretch", hide_index=True,
                        height=int((len(preview_orders) + 1) * 29 + 3), row_height=29)

                if st.button("✅ Apply imported configuration", type="primary", width="stretch", key="apply_excel_config"):
                    st.session_state.areas = preview_areas
                    st.session_state.order_types = preview_orders
                    st.session_state._saved_sig = _config_signature()
                    if db.save_all(preview_areas, preview_orders):
                        where = "the database" if db.storage_mode() == "database" else "local storage"
                        st.success("Configuration loaded and saved to " + where + ".")
                    else:
                        st.warning("Configuration loaded, but saving failed — applied for this session only.")
                    st.rerun()

            except Exception as e:
                st.error("❌ Could not parse Excel file: " + str(e))
                st.caption(
                    "Common causes: the 'Areas' or 'Order Types' sheet was renamed/removed, "
                    "a required column header was changed, or a percentage cell has text instead of a number."
                )
                with st.expander("🔍 Sheet names found in uploaded file"):
                    try:
                        from openpyxl import load_workbook as _lw
                        wb_debug = _lw(io.BytesIO(uploaded.getvalue()))
                        st.write(wb_debug.sheetnames)
                    except Exception:
                        st.caption("Could not read the uploaded file at all — it may not be a valid .xlsx file.")


def _render_db_controls():
    st.subheader("💾 Data & storage")
    mode = db.storage_mode()
    if mode == "database":
        st.success(
            "🟢 **Connected to a database.** Every Save persists across sessions, "
            "devices, and app restarts."
        )
    else:
        st.info(
            "🟢 **Your changes are saved automatically** and persist across page "
            "refreshes. For durable multi-user storage that also survives server "
            "restarts, you can optionally connect a free Supabase database below."
        )

    st.markdown("**Start fresh**")
    sc1, sc2 = st.columns(2)
    with sc1:
        if st.button("📥 Load calibrated example data", width="stretch",
                     help="Fill the app with the built-in Sojo example config (5 areas, 10 order types)."):
            st.session_state.areas       = [copy.deepcopy(a) for a in DEFAULT_AREAS]
            st.session_state.order_types = [copy.deepcopy(o) for o in DEFAULT_ORDER_TYPES]
            st.session_state._saved_sig  = _config_signature()
            db.save_all(st.session_state.areas, st.session_state.order_types)
            st.success("Loaded and saved the example data.")
            st.rerun()
    with sc2:
        if st.button("🗑️ Clear all data (start empty)", width="stretch",
                     help="Remove every area and order type. The app becomes a blank slate you populate by uploading a data file."):
            st.session_state._confirm_clear = True

    if st.session_state.get("_confirm_clear"):
        st.warning("**Remove ALL data?** Download a data file first if you want a backup.")
        xc1, xc2 = st.columns(2)
        if xc1.button("Yes, clear everything", type="primary", width="stretch", key="confirm_clear_yes"):
            db.clear_all()
            st.session_state.areas = []
            st.session_state.order_types = []
            st.session_state._saved_sig = _config_signature()
            st.session_state.pop("_confirm_clear", None)
            st.success("All data cleared. Upload a data file in Import / Export to begin.")
            st.rerun()
        if xc2.button("Cancel", width="stretch", key="confirm_clear_no"):
            st.session_state.pop("_confirm_clear", None)
            st.rerun()

    st.markdown("---")
    dbc1, dbc2 = st.columns(2)
    with dbc1:
        if st.button("🔄 Reload saved values", width="stretch",
                     help="Discard unsaved edits and reload the last saved configuration."):
            _areas, _orders = db.load_all()
            st.session_state.areas = _areas
            st.session_state.order_types = _orders
            st.session_state._saved_sig = _config_signature()
            st.session_state.pop("_confirm_reset", None)
            st.success("Reloaded the saved configuration.")
            st.rerun()
    with dbc2:
        if st.button("↩️ Reset to factory defaults", width="stretch",
                     help="Replace everything with the built-in calibrated defaults."):
            st.session_state._confirm_reset = True

    # two-step confirmation so nothing is wiped by a stray click
    if st.session_state.get("_confirm_reset"):
        st.warning("**Reset everything to the built-in defaults?** This discards your current values.")
        cc1, cc2 = st.columns(2)
        if cc1.button("Yes, reset to defaults", type="primary", width="stretch", key="confirm_reset_yes"):
            db.reset_to_defaults()
            st.session_state.areas = [copy.deepcopy(a) for a in DEFAULT_AREAS]
            st.session_state.order_types = [copy.deepcopy(o) for o in DEFAULT_ORDER_TYPES]
            st.session_state._saved_sig = _config_signature()
            st.session_state.pop("_confirm_reset", None)
            st.success("Reset to the built-in defaults.")
            st.rerun()
        if cc2.button("Cancel", width="stretch", key="confirm_reset_no"):
            st.session_state.pop("_confirm_reset", None)
            st.rerun()

    with st.expander("🔌 Connect a database (optional) — Supabase setup & diagnostic",
                     expanded=False):
        st.caption(
            "Optional. Connect Supabase (free, ~5 min — see README) for storage that "
            "persists across server restarts and is shared across everyone using the app."
        )
        diag = db.diagnose()
        steps = [
            ("Secrets section [supabase] found", diag["secrets_section_found"]),
            ("url value found",                  diag["url_found"]),
            ("key value found",                  diag["key_found"]),
            ("Client created",                   diag["client_created"]),
            ("Test query to database succeeded", diag["query_succeeded"]),
        ]
        for label, ok in steps:
            if ok:
                st.markdown("✅ " + label)
            else:
                st.markdown("⬜ " + label)
                break
        if diag["url_preview"]:
            st.caption("URL detected: " + diag["url_preview"])
        if diag["error"] and diag["secrets_section_found"]:
            st.error(diag["error"])
        st.caption(
            "After changing Secrets on Streamlit Cloud, wait ~20–30 sec and reboot "
            "the app (⋮ menu → Reboot app) to pick them up."
        )


if page == "Settings":
    st.title("Settings")
    render_last_updated()
    st.caption("Pick a section below. Edit values, then hit **Save & recalculate** — "
               "your changes are saved automatically and persist across refreshes.")

    save_status = st.empty()

    tab_areas, tab_orders, tab_kitting, tab_io = st.tabs([
        "Storage Areas",
        "Order Types",
        "Kitting",
        "Import / Export",
    ])

    with tab_areas:
        area_updates = _render_area_settings()

    with tab_orders:
        order_updates, orders_valid = _render_order_settings()

    with tab_kitting:
        kit_updates, kit_valid = _render_kitting_settings()

    with tab_io:
        _render_excel_io()

    # merge kitting fields into each order type's update dict
    for oid, ku in kit_updates.items():
        if oid in order_updates:
            order_updates[oid].update(ku)
        else:
            order_updates[oid] = ku

    # ── auto-save: apply edits from all editors every run, persist when valid ──
    _apply_updates(area_updates, order_updates)
    status = _autosave_if_changed(valid=orders_valid and kit_valid)
    if status == "saved":
        where = "database" if db.storage_mode() == "database" else "this server"
        save_status.success("✅ All changes saved automatically to " + where + ".")
    elif status == "blocked":
        save_status.warning("⚠️ Changes are **not saved** — fix the order-type splits (each pair must total 100%).")
    else:  # unchanged
        save_status.caption("✓ All changes saved. Edits save automatically as you type — no button needed.")


# ═══════════════════════════════════════════════════════════════════════════════
#  ANALYSIS PAGE
# ═══════════════════════════════════════════════════════════════════════════════

elif page == "Analysis":
    st.title("Analysis")
    render_last_updated()

    if not st.session_state.get("_intro_dismissed"):
        with st.container(border=True):
            st.markdown("#### Welcome to the Warehouse Capacity Planner")
            st.markdown(
                "- **Analysis** (this page) — how full each storage area is today. "
                "Drag the **Volume multiplier** to simulate order growth and find bottlenecks.\n"
                "- **Material flow** — how product moves from paper/consumables through the "
                "customer zones to final packout.\n"
                "- **Settings** — edit rack & box sizes, order volumes, and splits. "
                "Changes **save automatically** — no Save button needed.\n"
                "- Use **⬇️ Download report** below to export the current numbers to Excel."
            )
            if st.button("Got it — hide this", key="dismiss_intro"):
                st.session_state._intro_dismissed = True
                st.rerun()

    if not st.session_state.get("areas"):
        st.info(
            "**No data yet.** Go to **Settings \u2192 Import / Export** and upload "
            "your data file. The analysis will appear here once data is loaded."
        )
        st.stop()

    multiplier = st.slider(
        "Volume multiplier", min_value=1.0, max_value=50.0,
        value=1.0, step=0.5, format="x%.1f",
        help="Scale every order type's daily volume to model growth. x1.0 = today; "
             "x2.0 = double the volume. Drag up to x50 to stress-test. "
             "Watch which areas turn red first.")

    engine = get_engine()
    snap   = engine.snapshot(multiplier=multiplier)

    # ── Order volume tiles — what volume we're working with, and where it lands ──
    st.markdown("**Daily order volume** at x" + ("%.1f" % multiplier)
                + (" (today)" if multiplier == 1.0 else ""))
    zone_short = {"600": "Paper", "400": "Consumables", "300": "Cust 1",
                  "200": "Cust 2", "100": "Packout"}
    ocols = st.columns(len(engine.order_types))
    for col, ot in zip(ocols, engine.order_types):
        daily = int(round(ot.daily_volume * multiplier))
        units = int(round(ot.total_units() * multiplier)) if hasattr(ot, "total_units") else daily * ot.avg_units_per_order
        # boxes this order type puts into each area, biggest first (all areas,
        # so lower-volume areas like Paper are never hidden)
        contrib = sorted(
            [(a, ot.boxes_in_area(a, multiplier)) for a in engine.areas],
            key=lambda t: t[1], reverse=True)
        contrib = [(a, b) for a, b in contrib if b > 0]
        chips = "".join(
            "<span style='display:inline-block;background:#1e2235;border:1px solid #2e3250;"
            "border-radius:6px;padding:2px 8px;margin:2px 4px 0 0;font-size:12px;color:#c7d2fe'>"
            + zone_short.get(a.zone, a.zone) + " · " + str(int(round(b))) + "</span>"
            for a, b in contrib)
        col.markdown(
            "<div style='padding:12px 14px;border:1px solid #2e3250;border-radius:10px;"
            "border-top:3px solid #4f6ef7;background:#141728;'>"
            "<div style='font-size:13px;color:#9ca3af'>" + ot.name + "</div>"
            "<div style='font-size:26px;font-weight:700;color:#e8eaf6;line-height:1.2'>"
            + format(daily, ",") + "</div>"
            "<div style='font-size:12px;color:#6b7280;margin-bottom:6px'>orders/day · "
            + format(units, ",") + " units</div>"
            "<div style='font-size:11px;color:#6b7280'>Boxes into:</div>"
            "<div>" + chips + "</div>"
            "</div>", unsafe_allow_html=True)
    st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)

    k1, k2, k3, k4, k5, k6 = st.columns(6)
    k1.metric("Capacity (boxes)", str(snap.total_capacity_boxes),
              help="Total boxes/pallets all storage areas can hold when full.")
    k2.metric("Capacity (units)", str(snap.total_capacity_units),
              help="Total individual units across all areas (boxes × units per box).")
    k3.metric("Load (boxes)",     str(int(snap.total_load_boxes)),
              help="Boxes needed for the current daily order volume at this multiplier.")
    k4.metric("Overall util.",    str(round(snap.overall_utilization, 1)) + "%",
              help="Total load ÷ total capacity. Above 85% is tight; 100%+ is over capacity.")
    k5.metric("Over capacity",    len(snap.bottlenecks),
              help="Number of areas above 100% utilisation — these are your bottlenecks.")
    k6.metric("Near limit",       len(snap.warnings),
              help="Number of areas between 85% and 100% — approaching capacity.")

    for a in snap.bottlenecks:
        st.error("BOTTLENECK — " + a.area.name + " at " + str(round(a.utilization_pct, 1)) + "%")
    for a in snap.warnings:
        st.warning("WARNING — " + a.area.name + " at " + str(round(a.utilization_pct, 1)) + "%")
    if not snap.bottlenecks and not snap.warnings:
        st.success("All areas within capacity at x" + str(multiplier))

    _report_ts = datetime.now().strftime("%Y%m%d_%H%M")
    st.download_button(
        "⬇️ Download report (Excel)",
        data=analysis_report_excel_bytes(engine, multiplier),
        file_name="warehouse_analysis_x" + format(multiplier, ".1f") + "_" + _report_ts + ".xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="Export the current analysis (summary, per-area, per-zone, growth-to-capacity) "
             "at this multiplier as a formatted Excel workbook to share.",
    )

    st.markdown("---")
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "Overview", "Area detail", "Growth table", "Bottlenecks", "BOM breakdown"])

    with tab1:
        st.subheader("Area utilization")
        area_names = [a.area.name for a in snap.areas]
        utils      = [round(a.utilization_pct, 1) for a in snap.areas]
        colors     = [status_color(p) for p in utils]

        fig = go.Figure(go.Bar(
            y=area_names, x=utils, orientation="h",
            marker_color=colors,
            text=[str(p) + "%" for p in utils], textposition="outside",
            textfont=dict(size=18, color="#e8eaf6"),
            customdata=list(zip(
                [round(a.load_boxes, 0) for a in snap.areas],
                [a.capacity_boxes       for a in snap.areas],
                [round(a.load_units, 0) for a in snap.areas],
                [a.capacity_units       for a in snap.areas],
                [a.area.units_per_box   for a in snap.areas],
            )),
            hovertemplate=(
                "<b>%{y}</b><br>Util: %{x:.1f}%<br>"
                "Boxes: %{customdata[0]:,.0f} / %{customdata[1]:,}<br>"
                "Units: %{customdata[2]:,.0f} / %{customdata[3]:,}<br>"
                "Units/box: %{customdata[4]:.0f}<extra></extra>"
            ),
        ))
        fig.add_vline(x=85,  line_dash="dot", line_color="#f59e0b", annotation_text="85%")
        fig.add_vline(x=100, line_dash="dot", line_color="#ef4444", annotation_text="100%")
        fig.update_layout(
            xaxis=dict(title="Utilization %", range=[0, 115],
                       title_font=dict(size=15), tickfont=dict(size=14)),
            yaxis=dict(autorange="reversed", tickfont=dict(size=16)),
            height=430, margin=dict(l=10, r=70, t=24, b=44),
            bargap=0.28,
            plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
        st.plotly_chart(fig, width="stretch")

        st.subheader("Zone summary")
        zone_df = engine.zone_summary(multiplier)
        zcols = st.columns(len(zone_df))
        for i, (_, row) in enumerate(zone_df.iterrows()):
            pct = row["utilization_pct"]
            zc  = ZONE_COLORS.get(row["zone_code"], "#888")
            with zcols[i]:
                st.markdown(
                    "<div style='text-align:center;padding:18px 14px;border:1px solid #2e3250;"
                    "border-radius:12px;border-left:6px solid " + zc + ";background:#141728;'>"
                    "<div style='font-size:13px;color:#6b7280'>Zone " + str(row["zone_code"]) + "</div>"
                    "<div style='font-weight:600;font-size:16px;color:#e8eaf6'>" + str(row["zone_name"]) + "</div>"
                    "<div style='font-size:38px;font-weight:800;color:" + status_color(pct) + ";line-height:1.3'>" + str(pct) + "%</div>"
                    "<div style='font-size:13px;color:#6b7280'>" + str(row["capacity_boxes"]) + " box capacity</div>"
                    "</div>", unsafe_allow_html=True)

        st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)

        rows = []
        for a in snap.areas:
            of_note = ""
            if a.overflow_capacity_boxes > 0:
                fb = a.fill or {}
                using = "main"
                if fb.get("of2", 0) > 0: using = "overflow 2"
                elif fb.get("of1", 0) > 0: using = "overflow 1"
                of_note = "+" + str(a.overflow_capacity_boxes) + " (" + using + ")"
            rows.append({
                "Area":         a.area.name,
                "Zone":         a.area.zone,
                "Area volume (ft³)": round(a.area.volume_cuft, 1),
                "Box L×W×H (ft)": fmt_dims_ft(a.area.box_length_cuft, a.area.box_depth_cuft, a.area.box_height_cuft),
                "Box vol (ft³)": round(a.area.avg_box_size_cuft, 3),
                "Units/box":    int(a.area.units_per_box),
                "Load (boxes)": str(int(a.load_boxes)),
                "Cap (boxes)":  str(a.capacity_boxes),
                "Overflow":     of_note or "—",
                "Total cap":    str(a.total_capacity_boxes),
                "Load (units)": str(int(a.load_units)),
                "Cap (units)":  str(a.capacity_units),
                "Util %":       str(round(a.utilization_pct, 1)) + "%",
                "Status":       status_label_for(a),
            })
        st.dataframe(pd.DataFrame(rows), width="stretch", hide_index=True)

    with tab2:
        st.subheader("Area detail — order contributions")
        for a in snap.areas:
            cap_tag = ""
            with st.expander(
                a.area.name + " — " + str(round(a.utilization_pct, 1)) + "%  |  "
                + str(int(a.load_boxes)) + "/" + str(a.capacity_boxes) + " boxes" + cap_tag + "  |  "
                + str(int(a.load_units)) + "/" + str(a.capacity_units) + " units  |  "
                + str(int(a.area.units_per_box)) + " u/box  " + status_label(a.utilization_pct),
                expanded=False):
                st.progress(min(a.utilization_pct / 100, 1.0))
                if a.contributing_orders:
                    rows = []
                    for oid, boxes in sorted(
                        a.contributing_orders.items(), key=lambda x: x[1], reverse=True):
                        rows.append({
                            "Order":     oid,
                            "Boxes":     str(round(boxes, 1)),
                            "Units":     str(int(boxes * a.area.units_per_box)),
                            "% of area": str(round(boxes / a.load_boxes * 100, 1)) + "%" if a.load_boxes else "—",
                        })
                    st.dataframe(pd.DataFrame(rows), width="stretch", hide_index=True)
                else:
                    st.caption("No orders currently load this area.")

    with tab3:
        st.subheader("Growth table")
        df = engine.growth_table(max_multiplier=50.0, steps=20)
        pivot = df.pivot_table(
            index="multiplier", columns="area",
            values="utilization_pct", aggfunc="first").reset_index()
        area_names_list = [a.name for a in engine.areas]

        fig3 = go.Figure(data=go.Heatmap(
            z=[[row.get(n, 0) for n in area_names_list] for _, row in pivot.iterrows()],
            x=area_names_list,
            y=["x" + str(round(row["multiplier"], 1)) for _, row in pivot.iterrows()],
            colorscale=[[0,"#1a3a2a"],[0.70,"#22c55e"],[0.85,"#f59e0b"],[1,"#ef4444"]],
            zmin=0, zmax=110,
            text=[[str(int(row.get(n, 0))) + "%" for n in area_names_list] for _, row in pivot.iterrows()],
            texttemplate="%{text}",
            hovertemplate="Area: %{x}<br>Mult: %{y}<br>Util: %{z:.1f}%<extra></extra>",
        ))
        fig3.update_layout(height=360, margin=dict(l=10,r=10,t=20,b=10),
            plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
            xaxis=dict(tickangle=-30))
        st.plotly_chart(fig3, width="stretch")

        def color_cell(val):
            try:
                v = float(str(val).replace("%", ""))
                if v >= 100: return "background-color:#7f1d1d;color:#fca5a5"
                if v >= 85:  return "background-color:#78350f;color:#fde68a"
                if v >= 70:  return "background-color:#431407;color:#fed7aa"
                return "background-color:#052e16;color:#86efac"
            except:
                return ""

        dp = pivot.copy()
        dp["multiplier"] = dp["multiplier"].apply(lambda x: "x" + str(round(x, 1)))
        dp = dp.rename(columns={"multiplier": "Mult."})
        for col in area_names_list:
            if col in dp.columns:
                dp[col] = dp[col].apply(lambda x: str(int(x)) + "%")
        st.dataframe(dp.style.map(color_cell, subset=area_names_list),
                     width="stretch", hide_index=True)

    with tab4:
        st.subheader("Bottleneck sequence")
        seq_100 = engine.bottleneck_sequence(threshold_pct=100.0, max_mult=50.0)
        seq_85  = engine.bottleneck_sequence(threshold_pct=85.0,  max_mult=50.0)

        if not seq_100:
            st.success("No areas hit 100% within x20 volume.")
        else:
            for rank, (mult, area_name, _) in enumerate(seq_100, 1):
                icon = "🔴" if rank == 1 else "🟡" if rank == 2 else "🟠"
                note = " — Address first" if rank == 1 else ""
                st.markdown(
                    "#" + str(rank) + " " + icon + " **" + area_name
                    + "** hits capacity at **x" + str(mult) + "**" + note)

        st.markdown("---")
        st.subheader("85% warning threshold")
        if seq_85:
            st.dataframe(
                pd.DataFrame([{"Area": n, "Reaches 85% at": "x" + str(m)} for m, n, _ in seq_85]),
                width="stretch", hide_index=True)
            fig4 = go.Figure()
            for mult, name, _ in reversed(seq_85):
                fig4.add_trace(go.Bar(
                    y=[name], x=[mult], orientation="h",
                    marker_color=status_color(85),
                    text=["x" + str(mult)], textposition="outside",
                    showlegend=False))
            fig4.add_vline(x=multiplier, line_dash="dash", line_color="#4f6ef7",
                           annotation_text="Current x" + str(multiplier))
            fig4.update_layout(
                xaxis=dict(title="Multiplier", range=[0, 22]),
                yaxis=dict(autorange="reversed"),
                height=220, margin=dict(l=10,r=60,t=40,b=40),
                plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
                barmode="overlay")
            st.plotly_chart(fig4, width="stretch")
        else:
            st.success("No areas reach 85% within x20 volume.")

    with tab5:
        st.subheader("BOM breakdown")
        bom_df = engine.order_bom_summary(multiplier)
        if not bom_df.empty:
            fig5 = px.bar(
                bom_df, x="order", y="boxes", color="zone_name", barmode="group",
                labels={"boxes":"Boxes/day","order":"Order type","zone_name":"Zone"},
                title="Daily boxes by zone and order type",
                color_discrete_map={
                    ZONE_NAMES["600"]: ZONE_COLORS["600"],
                    ZONE_NAMES["400"]: ZONE_COLORS["400"],
                    ZONE_NAMES["300"]: ZONE_COLORS["300"],
                    ZONE_NAMES["200"]: ZONE_COLORS["200"],
                })
            fig5.update_layout(height=250, margin=dict(l=10,r=10,t=40,b=40),
                plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig5, width="stretch")

            disp = bom_df[["order_name","zone_name","pct_of_total","units","units_per_box","boxes"]].copy()
            disp.columns = ["Order","Zone","% of total","Units/day","Units per box","Boxes/day"]
            st.dataframe(disp, width="stretch", hide_index=True)
